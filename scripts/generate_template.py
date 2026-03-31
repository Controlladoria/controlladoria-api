"""
Generate the financial template spreadsheet for customers.
Creates modelo-financeiro.xlsx with two sheets: Recebimentos + Pagamentos.

Usage: python scripts/generate_template.py
Output: frontend/public/templates/modelo-financeiro.xlsx
"""

import sys
from pathlib import Path

# Add parent dir to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


from accounting.categories import DRE_CATEGORIES

def get_category_display_names() -> list[str]:
    """Get all 52 DRE category display names for the dropdown."""
    names = sorted(set(cat.get("display_name", k) for k, cat in DRE_CATEGORIES.items()))
    return names


def generate_template():
    wb = openpyxl.Workbook()

    # Styles
    green_fill = PatternFill(start_color="095A5E", end_color="095A5E", fill_type="solid")
    red_fill = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
    header_font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="Inter", size=10)
    date_font = Font(name="Inter", size=10)
    thin_border = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD"),
    )

    # Type validation (shared)
    type_list = '"Receita,Despesa,Custo,Investimento,Perda"'
    type_validation = DataValidation(type="list", formula1=type_list, allow_blank=True)
    type_validation.error = "Selecione um tipo válido"
    type_validation.errorTitle = "Tipo inválido"
    type_validation.prompt = "Selecione o tipo de lançamento"
    type_validation.promptTitle = "Tipo"

    # Category validation
    try:
        cat_names = get_category_display_names()
        # Excel data validation list has a ~255 char limit for formula1 string
        # We'll create a hidden sheet with the category list
        cat_sheet = wb.create_sheet("_Categorias")
        for i, name in enumerate(cat_names, 1):
            cat_sheet.cell(row=i, column=1, value=name)
        cat_sheet.sheet_state = "hidden"
        cat_formula = f"_Categorias!$A$1:$A${len(cat_names)}"
        cat_validation = DataValidation(type="list", formula1=cat_formula, allow_blank=True)
        cat_validation.error = "Selecione uma categoria válida"
        cat_validation.errorTitle = "Categoria inválida"
        cat_validation.prompt = "Selecione a categoria contábil"
        cat_validation.promptTitle = "Categoria"
    except Exception:
        cat_validation = None

    # ================================================================
    # Sheet 1: Recebimentos (Receivables)
    # ================================================================
    ws_rec = wb.active
    ws_rec.title = "Recebimentos"

    rec_headers = ["Data", "Descrição", "Nº Documento", "Cliente", "Valor", "Tipo", "Categoria", "Observações"]
    rec_widths = [14, 40, 16, 35, 16, 16, 30, 30]

    for col_idx, (header, width) in enumerate(zip(rec_headers, rec_widths), 1):
        cell = ws_rec.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws_rec.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws_rec.freeze_panes = "A2"

    # Add type validation to Tipo column (F)
    type_val_rec = DataValidation(type="list", formula1=type_list, allow_blank=True)
    type_val_rec.prompt = "Selecione o tipo"
    type_val_rec.promptTitle = "Tipo"
    ws_rec.add_data_validation(type_val_rec)
    type_val_rec.add(f"F2:F1000")

    # Add category validation to Categoria column (G)
    if cat_validation:
        cat_val_rec = DataValidation(type="list", formula1=cat_formula, allow_blank=True)
        cat_val_rec.prompt = "Selecione a categoria"
        cat_val_rec.promptTitle = "Categoria"
        ws_rec.add_data_validation(cat_val_rec)
        cat_val_rec.add(f"G2:G1000")

    # Example rows
    # Use actual display names from the category system
    cat_display = {k: v.get("display_name", k) for k, v in DRE_CATEGORIES.items()}
    rec_examples = [
        ["15/10/2025", "Medição #3 - Obra Centro", "NF 888", "Prefeitura Municipal", 73034.75, "Receita", cat_display.get("receita_servicos", "receita_servicos"), "Transferência bancária"],
        ["20/10/2025", "Consultoria Financeira Out/25", "NF 892", "Empresa ABC Ltda", 15000.00, "Receita", cat_display.get("receita_servicos", "receita_servicos"), ""],
        ["28/10/2025", "Venda de Material Excedente", "NF 895", "Construções Silva", 5200.00, "Receita", cat_display.get("receita_vendas_produtos", "receita_vendas_produtos"), "PIX recebido"],
    ]

    for row_idx, example in enumerate(rec_examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws_rec.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx == 5:  # Valor
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 1:  # Data
                cell.alignment = Alignment(horizontal="center")

    # ================================================================
    # Sheet 2: Pagamentos (Payables)
    # ================================================================
    ws_pag = wb.create_sheet("Pagamentos")

    pag_headers = ["Data", "Descrição", "Nº Documento", "Fornecedor", "Valor", "Tipo", "Categoria", "Centro de Custo", "Observações"]
    pag_widths = [14, 40, 16, 35, 16, 16, 30, 20, 30]

    for col_idx, (header, width) in enumerate(zip(pag_headers, pag_widths), 1):
        cell = ws_pag.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = red_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws_pag.column_dimensions[get_column_letter(col_idx)].width = width

    ws_pag.freeze_panes = "A2"

    # Type validation for Pagamentos (F column)
    type_val_pag = DataValidation(type="list", formula1=type_list, allow_blank=True)
    type_val_pag.prompt = "Selecione o tipo"
    type_val_pag.promptTitle = "Tipo"
    ws_pag.add_data_validation(type_val_pag)
    type_val_pag.add(f"F2:F1000")

    # Category validation for Pagamentos (G column)
    if cat_validation:
        cat_val_pag = DataValidation(type="list", formula1=cat_formula, allow_blank=True)
        cat_val_pag.prompt = "Selecione a categoria"
        cat_val_pag.promptTitle = "Categoria"
        ws_pag.add_data_validation(cat_val_pag)
        cat_val_pag.add(f"G2:G1000")

    # Example rows
    pag_examples = [
        ["02/10/2025", "Material de Construção - Cimento", "NF 4521", "Materiais ABC", 3500.00, "Custo", cat_display.get("materia_prima", "materia_prima"), "Obra Centro", "PIX"],
        ["10/10/2025", "Aluguel Escritório Out/25", "BOL 1025", "Imobiliária XYZ", 4200.00, "Despesa", cat_display.get("aluguel", "aluguel"), "Administrativo", "Boleto"],
        ["15/10/2025", "Compra Betoneira Modelo X", "NF 7890", "Equipamentos Pesados", 25000.00, "Investimento", cat_display.get("maquinas_equipamentos", "maquinas_equipamentos"), "Obra Centro", "Financiamento 12x"],
    ]

    for row_idx, example in enumerate(pag_examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws_pag.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx == 5:  # Valor
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 1:  # Data
                cell.alignment = Alignment(horizontal="center")

    # Save
    output_path = Path(__file__).resolve().parent.parent / "frontend" / "public" / "templates" / "modelo-financeiro.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Template saved to: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
    if cat_validation:
        print(f"  Categories: {len(cat_names)} loaded for dropdown")


if __name__ == "__main__":
    generate_template()
