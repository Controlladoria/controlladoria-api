"""
Balance Sheet Export Functions - V2 (Template-Based)
Exports Balance Sheet to PDF, Excel, and CSV formats with Brazilian formatting

V2 Changes (matching ControlladorIA_Template_V0):
- Side-by-side layout: ATIVO (left) | PASSIVO+PL (right) in Excel
- 4-group asset structure: Circulante, Não Circulante, Imobilizado, Intangível
- Green color scheme matching template
- Year column headers
"""

import csv
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Optional

from .balance_sheet_calculator import BalanceSheet


def format_brl(value: Decimal) -> str:
    """Format decimal as Brazilian currency R$ 1.234,56"""
    abs_value = abs(value)
    formatted = f"{abs_value:,.2f}"
    # Convert to Brazilian format: 1,234.56 -> 1.234,56
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    if value < 0:
        return f"(R$ {formatted})"
    return f"R$ {formatted}"


# =========================================
# PDF EXPORT (V2)
# =========================================


def export_balance_sheet_to_pdf(balance_sheet: BalanceSheet, logo_bytes: bytes = None) -> bytes:
    """
    Export Balance Sheet to PDF - V2 layout matching template

    Args:
        balance_sheet: BalanceSheet object
        logo_bytes: Optional PNG/JPEG bytes for the header logo

    Returns:
        PDF file as bytes
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import Image as RLImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        raise ImportError(
            "reportlab is required for PDF export. Install with: pip install reportlab"
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=10 * mm, bottomMargin=10 * mm
    )

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1B5E20"),
        spaceAfter=6,
        alignment=TA_CENTER,
    )

    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#666666"),
        spaceAfter=12,
        alignment=TA_CENTER,
    )

    story = []

    # Logo header
    if logo_bytes:
        try:
            img = RLImage(BytesIO(logo_bytes), width=5 * cm, height=2 * cm, kind="proportional")
            img.hAlign = "CENTER"
            story.append(img)
            story.append(Spacer(1, 3 * mm))
        except Exception:
            pass

    # Header
    story.append(Paragraph("BALANÇO GERENCIAL", title_style))

    if balance_sheet.company_name:
        story.append(Paragraph(balance_sheet.company_name, subtitle_style))
    if balance_sheet.cnpj:
        story.append(Paragraph(f"CNPJ: {balance_sheet.cnpj}", subtitle_style))

    story.append(
        Paragraph(
            f"Data de Referência: {balance_sheet.reference_date.strftime('%d/%m/%Y')}",
            subtitle_style,
        )
    )
    story.append(Spacer(1, 12))

    # ================================================================
    # ATIVO table
    # ================================================================
    def _build_asset_data():
        data = [
            [
                Paragraph("<b>ATIVO</b>", styles["Normal"]),
                Paragraph(f"<b>{balance_sheet.reference_date.year}</b>", styles["Normal"]),
            ]
        ]

        # Group 1: Ativo Circulante
        data.append([Paragraph("<b>Ativo Circulante</b>", styles["Normal"]),
                      Paragraph(f"<b>{format_brl(balance_sheet.ativo_circulante)}</b>", styles["Normal"])])
        for line in balance_sheet.asset_lines:
            data.append([Paragraph(f"&nbsp;&nbsp;{line.name}", styles["Normal"]),
                          Paragraph(format_brl(line.balance), styles["Normal"])])

        # Group 2: Ativo Não Circulante
        data.append([Paragraph("<b>Ativo Não Circulante</b>", styles["Normal"]),
                      Paragraph(f"<b>{format_brl(balance_sheet.ativo_nao_circulante)}</b>", styles["Normal"])])
        for line in balance_sheet.asset_noncurrent_lines:
            data.append([Paragraph(f"&nbsp;&nbsp;{line.name}", styles["Normal"]),
                          Paragraph(format_brl(line.balance), styles["Normal"])])

        # Group 3: Imobilizado
        data.append([Paragraph("<b>Imobilizado</b>", styles["Normal"]),
                      Paragraph(f"<b>{format_brl(balance_sheet.imobilizado)}</b>", styles["Normal"])])
        for line in balance_sheet.imobilizado_lines:
            data.append([Paragraph(f"&nbsp;&nbsp;{line.name}", styles["Normal"]),
                          Paragraph(format_brl(line.balance), styles["Normal"])])

        # Group 4: Intangível
        data.append([Paragraph("<b>Intangível</b>", styles["Normal"]),
                      Paragraph(f"<b>{format_brl(balance_sheet.intangivel)}</b>", styles["Normal"])])
        for line in balance_sheet.intangivel_lines:
            data.append([Paragraph(f"&nbsp;&nbsp;{line.name}", styles["Normal"]),
                          Paragraph(format_brl(line.balance), styles["Normal"])])

        # TOTAL ATIVO
        data.append([Paragraph("<b>TOTAL DO ATIVO</b>", styles["Normal"]),
                      Paragraph(f"<b>{format_brl(balance_sheet.total_ativo)}</b>", styles["Normal"])])
        return data

    asset_data = _build_asset_data()
    asset_table = Table(asset_data, colWidths=[130 * mm, 50 * mm])
    asset_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("LINEABOVE", (0, -1), (-1, -1), 2, colors.black),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#C8E6C9")),
    ]))

    story.append(asset_table)
    story.append(Spacer(1, 12))

    # ================================================================
    # PASSIVO + PL table
    # ================================================================
    def _build_passivo_data():
        data = [
            [
                Paragraph("<b>PASSIVO + PL</b>", styles["Normal"]),
                Paragraph(f"<b>{balance_sheet.reference_date.year}</b>", styles["Normal"]),
            ]
        ]

        # Passivo Circulante
        data.append([Paragraph("<b>Passivo Circulante</b>", styles["Normal"]),
                      Paragraph(f"<b>{format_brl(balance_sheet.passivo_circulante)}</b>", styles["Normal"])])
        for line in balance_sheet.liability_lines:
            if line.code.startswith("2.01"):
                data.append([Paragraph(f"&nbsp;&nbsp;{line.name}", styles["Normal"]),
                              Paragraph(format_brl(line.balance), styles["Normal"])])

        # Passivo Não Circulante
        data.append([Paragraph("<b>Passivo Não Circulante</b>", styles["Normal"]),
                      Paragraph(f"<b>{format_brl(balance_sheet.passivo_nao_circulante)}</b>", styles["Normal"])])
        for line in balance_sheet.liability_lines + balance_sheet.liability_noncurrent_lines:
            if line.code.startswith("2.02"):
                data.append([Paragraph(f"&nbsp;&nbsp;{line.name}", styles["Normal"]),
                              Paragraph(format_brl(line.balance), styles["Normal"])])

        # Patrimônio Líquido
        data.append([Paragraph("<b>Patrimônio Líquido</b>", styles["Normal"]),
                      Paragraph(f"<b>{format_brl(balance_sheet.patrimonio_liquido)}</b>", styles["Normal"])])
        for line in balance_sheet.equity_lines:
            data.append([Paragraph(f"&nbsp;&nbsp;{line.name}", styles["Normal"]),
                          Paragraph(format_brl(line.balance), styles["Normal"])])

        # TOTAL PASSIVO + PL
        total_passivo_pl = balance_sheet.total_passivo + balance_sheet.patrimonio_liquido
        data.append([Paragraph("<b>TOTAL DO PASSIVO + PL</b>", styles["Normal"]),
                      Paragraph(f"<b>{format_brl(total_passivo_pl)}</b>", styles["Normal"])])
        return data

    passivo_data = _build_passivo_data()
    passivo_table = Table(passivo_data, colWidths=[130 * mm, 50 * mm])
    passivo_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("LINEABOVE", (0, -1), (-1, -1), 2, colors.black),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#C8E6C9")),
    ]))

    story.append(passivo_table)
    story.append(Spacer(1, 12))

    # Balance check
    if not balance_sheet.is_balanced:
        from reportlab.lib.enums import TA_CENTER
        warning_style = ParagraphStyle("Warning", parent=styles["Normal"],
                                        fontSize=10, textColor=colors.red, alignment=TA_CENTER)
        story.append(Paragraph(
            f"ATENÇÃO: Balanço não equilibrado. Diferença: {format_brl(balance_sheet.balance_difference)}",
            warning_style,
        ))

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


# =========================================
# EXCEL EXPORT (V2 - Template-based side-by-side)
# =========================================


def export_balance_sheet_to_excel(balance_sheet: BalanceSheet, logo_bytes: bytes = None) -> bytes:
    """
    Export Balance Sheet to Excel matching ControlladorIA_Template_V0 layout.

    Layout (side-by-side):
    - B2: "BALANÇO GERENCIAL"
    - Left side (B,D,F): ATIVO with year columns
    - Right side (H,J): PASSIVO + PL with year columns
    - Green color scheme

    Args:
        balance_sheet: BalanceSheet object
        logo_bytes: Optional PNG/JPEG bytes for the header logo

    Returns:
        Excel file as bytes
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "BALANÇO GERENCIAL"

    # Logo header (floats over A1)
    if logo_bytes:
        try:
            from io import BytesIO as _BytesIO
            from openpyxl.drawing.image import Image as XLImage
            xl_img = XLImage(_BytesIO(logo_bytes))
            xl_img.width = 140
            xl_img.height = 50
            ws.add_image(xl_img, "A1")
        except Exception:
            pass

    # ================================================================
    # Styles
    # ================================================================
    green_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    light_green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    title_font = Font(name="Arial", size=14, bold=True, color="1B5E20")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Arial", size=11, bold=True)
    total_font = Font(name="Arial", size=11, bold=True)
    normal_font = Font(name="Arial", size=10)
    sub_font = Font(name="Arial", size=10, color="333333")

    thin_border = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD"),
    )

    year_str = str(balance_sheet.reference_date.year)

    # ================================================================
    # Title
    # ================================================================
    ws["B2"] = "BALANÇO GERENCIAL"
    ws["B2"].font = title_font
    ws.merge_cells("B2:F2")

    if balance_sheet.company_name:
        ws["B3"] = balance_sheet.company_name
        ws["B3"].font = Font(name="Arial", size=11, color="666666")

    # ================================================================
    # Column headers
    # ================================================================
    # ATIVO side (columns B, D)
    ws["D4"] = year_str
    ws["D4"].font = header_font
    ws["D4"].fill = green_fill
    ws["D4"].alignment = Alignment(horizontal="center")
    ws["D4"].border = thin_border

    # PASSIVO side (columns H, J)
    ws["J4"] = year_str
    ws["J4"].font = header_font
    ws["J4"].fill = green_fill
    ws["J4"].alignment = Alignment(horizontal="center")
    ws["J4"].border = thin_border

    # ================================================================
    # ATIVO side (left)
    # ================================================================
    row = 5

    # ATIVO header
    ws[f"B{row}"] = "ATIVO"
    ws[f"B{row}"].font = section_font
    ws[f"B{row}"].fill = light_green_fill
    ws[f"D{row}"] = float(balance_sheet.total_ativo)
    ws[f"D{row}"].font = total_font
    ws[f"D{row}"].fill = light_green_fill
    ws[f"D{row}"].number_format = '#,##0.00'
    ws[f"D{row}"].alignment = Alignment(horizontal="right")
    for col in ["B", "D"]:
        ws[f"{col}{row}"].border = thin_border
    row += 1

    # Skip a row
    row += 1  # row 7

    # Ativo Circulante
    ws[f"B{row}"] = "Ativo Circulante"
    ws[f"B{row}"].font = section_font
    ws[f"D{row}"] = float(balance_sheet.ativo_circulante)
    ws[f"D{row}"].font = total_font
    ws[f"D{row}"].number_format = '#,##0.00'
    ws[f"D{row}"].alignment = Alignment(horizontal="right")
    for col in ["B", "D"]:
        ws[f"{col}{row}"].border = thin_border
    row += 1

    # Ativo Circulante items (template: Caixa, Aplicações, Clientes, Estoques, Despesas Antecipadas)
    template_ac_items = [
        "Caixa e Equivalentes de Caixa",
        "Aplicações Financeiras (curto prazo)",
        "Clientes (Contas a Receber)",
        "Estoques",
        "Despesas Antecipadas e Outros",
    ]
    ac_values = {}
    for line in balance_sheet.asset_lines:
        ac_values[line.name] = float(line.balance)

    for item_name in template_ac_items:
        ws[f"B{row}"] = item_name
        ws[f"B{row}"].font = sub_font
        ws[f"D{row}"] = ac_values.get(item_name, 0)
        ws[f"D{row}"].number_format = '#,##0.00'
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        ws[f"D{row}"].font = normal_font
        for col in ["B", "D"]:
            ws[f"{col}{row}"].border = thin_border
        row += 1

    # Any asset lines not in template
    for line in balance_sheet.asset_lines:
        if line.name not in template_ac_items:
            ws[f"B{row}"] = line.name
            ws[f"B{row}"].font = sub_font
            ws[f"D{row}"] = float(line.balance)
            ws[f"D{row}"].number_format = '#,##0.00'
            ws[f"D{row}"].alignment = Alignment(horizontal="right")
            for col in ["B", "D"]:
                ws[f"{col}{row}"].border = thin_border
            row += 1

    row += 1  # blank row

    # Ativo Não Circulante
    ws[f"B{row}"] = "Ativo Não Circulante"
    ws[f"B{row}"].font = section_font
    ws[f"D{row}"] = float(balance_sheet.ativo_nao_circulante)
    ws[f"D{row}"].font = total_font
    ws[f"D{row}"].number_format = '#,##0.00'
    ws[f"D{row}"].alignment = Alignment(horizontal="right")
    for col in ["B", "D"]:
        ws[f"{col}{row}"].border = thin_border
    row += 1

    for line in balance_sheet.asset_noncurrent_lines:
        ws[f"B{row}"] = line.name
        ws[f"B{row}"].font = sub_font
        ws[f"D{row}"] = float(line.balance)
        ws[f"D{row}"].number_format = '#,##0.00'
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        for col in ["B", "D"]:
            ws[f"{col}{row}"].border = thin_border
        row += 1

    # Imobilizado
    ws[f"B{row}"] = "Imobilizado"
    ws[f"B{row}"].font = section_font
    ws[f"D{row}"] = float(balance_sheet.imobilizado)
    ws[f"D{row}"].font = total_font
    ws[f"D{row}"].number_format = '#,##0.00'
    ws[f"D{row}"].alignment = Alignment(horizontal="right")
    for col in ["B", "D"]:
        ws[f"{col}{row}"].border = thin_border
    row += 1

    for line in balance_sheet.imobilizado_lines:
        ws[f"B{row}"] = line.name
        ws[f"B{row}"].font = sub_font
        ws[f"D{row}"] = float(line.balance)
        ws[f"D{row}"].number_format = '#,##0.00'
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        for col in ["B", "D"]:
            ws[f"{col}{row}"].border = thin_border
        row += 1

    # Intangível
    ws[f"B{row}"] = "Intangível"
    ws[f"B{row}"].font = section_font
    ws[f"D{row}"] = float(balance_sheet.intangivel)
    ws[f"D{row}"].font = total_font
    ws[f"D{row}"].number_format = '#,##0.00'
    ws[f"D{row}"].alignment = Alignment(horizontal="right")
    for col in ["B", "D"]:
        ws[f"{col}{row}"].border = thin_border
    row += 1

    for line in balance_sheet.intangivel_lines:
        ws[f"B{row}"] = line.name
        ws[f"B{row}"].font = sub_font
        ws[f"D{row}"] = float(line.balance)
        ws[f"D{row}"].number_format = '#,##0.00'
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        for col in ["B", "D"]:
            ws[f"{col}{row}"].border = thin_border
        row += 1

    # ================================================================
    # PASSIVO + PL side (right, starting at row 5)
    # ================================================================
    prow = 5

    # PASSIVO + PL header
    ws[f"H{prow}"] = "PASSIVO  + PL"
    ws[f"H{prow}"].font = section_font
    ws[f"H{prow}"].fill = light_green_fill
    total_passivo_pl = balance_sheet.total_passivo + balance_sheet.patrimonio_liquido
    ws[f"J{prow}"] = float(total_passivo_pl)
    ws[f"J{prow}"].font = total_font
    ws[f"J{prow}"].fill = light_green_fill
    ws[f"J{prow}"].number_format = '#,##0.00'
    ws[f"J{prow}"].alignment = Alignment(horizontal="right")
    for col in ["H", "J"]:
        ws[f"{col}{prow}"].border = thin_border
    prow += 2  # skip row

    # Passivo Circulante
    ws[f"H{prow}"] = "Passivo Circulante"
    ws[f"H{prow}"].font = section_font
    ws[f"J{prow}"] = float(balance_sheet.passivo_circulante)
    ws[f"J{prow}"].font = total_font
    ws[f"J{prow}"].number_format = '#,##0.00'
    ws[f"J{prow}"].alignment = Alignment(horizontal="right")
    for col in ["H", "J"]:
        ws[f"{col}{prow}"].border = thin_border
    prow += 1

    template_pc_items = [
        "Fornecedores",
        "Empréstimos e Financiamentos (CP)",
        "Obrigações Trabalhistas e Sociais",
        "Obrigações Fiscais",
        "Provisões e Outros",
    ]
    pc_values = {}
    for line in balance_sheet.liability_lines:
        if line.code.startswith("2.01"):
            pc_values[line.name] = float(line.balance)

    for item_name in template_pc_items:
        ws[f"H{prow}"] = item_name
        ws[f"H{prow}"].font = sub_font
        ws[f"J{prow}"] = pc_values.get(item_name, 0)
        ws[f"J{prow}"].number_format = '#,##0.00'
        ws[f"J{prow}"].alignment = Alignment(horizontal="right")
        ws[f"J{prow}"].font = normal_font
        for col in ["H", "J"]:
            ws[f"{col}{prow}"].border = thin_border
        prow += 1

    prow += 1  # blank row

    # Passivo Não Circulante
    ws[f"H{prow}"] = "Passivo Não Circulante"
    ws[f"H{prow}"].font = section_font
    ws[f"J{prow}"] = float(balance_sheet.passivo_nao_circulante)
    ws[f"J{prow}"].font = total_font
    ws[f"J{prow}"].number_format = '#,##0.00'
    ws[f"J{prow}"].alignment = Alignment(horizontal="right")
    for col in ["H", "J"]:
        ws[f"{col}{prow}"].border = thin_border
    prow += 1

    template_pnc_items = [
        "Empréstimos e Financiamentos (LP)",
        "Provisões (LP)",
        "Passivos Fiscais Diferidos",
    ]
    pnc_values = {}
    for line in balance_sheet.liability_lines + balance_sheet.liability_noncurrent_lines:
        if line.code.startswith("2.02"):
            pnc_values[line.name] = float(line.balance)

    for item_name in template_pnc_items:
        ws[f"H{prow}"] = item_name
        ws[f"H{prow}"].font = sub_font
        ws[f"J{prow}"] = pnc_values.get(item_name, 0)
        ws[f"J{prow}"].number_format = '#,##0.00'
        ws[f"J{prow}"].alignment = Alignment(horizontal="right")
        ws[f"J{prow}"].font = normal_font
        for col in ["H", "J"]:
            ws[f"{col}{prow}"].border = thin_border
        prow += 1

    prow += 1  # blank row

    # Patrimônio Líquido
    ws[f"H{prow}"] = "Patrimônio Líquido"
    ws[f"H{prow}"].font = section_font
    ws[f"J{prow}"] = float(balance_sheet.patrimonio_liquido)
    ws[f"J{prow}"].font = total_font
    ws[f"J{prow}"].number_format = '#,##0.00'
    ws[f"J{prow}"].alignment = Alignment(horizontal="right")
    for col in ["H", "J"]:
        ws[f"{col}{prow}"].border = thin_border
    prow += 1

    # PL breakdown: use code-prefix matching to aggregate by family
    # Must match calculator.to_dict() groupings exactly:
    #   3.01 = Capital Social, 3.02+3.03 = Reservas, 3.04+3.05 = Lucros
    template_pl_items = [
        ("Capital Social", ["3.01"]),
        ("Reservas e Ajustes", ["3.02", "3.03"]),
        ("Lucro ou prejuízo do exercício", ["3.04", "3.05"]),
    ]
    for item_name, code_prefixes in template_pl_items:
        val = sum(float(l.balance) for l in balance_sheet.equity_lines if any(l.code.startswith(p) for p in code_prefixes))
        ws[f"H{prow}"] = item_name
        ws[f"H{prow}"].font = sub_font
        ws[f"J{prow}"] = val
        ws[f"J{prow}"].number_format = '#,##0.00'
        ws[f"J{prow}"].alignment = Alignment(horizontal="right")
        ws[f"J{prow}"].font = normal_font
        for col in ["H", "J"]:
            ws[f"{col}{prow}"].border = thin_border
        prow += 1

    # ================================================================
    # Column widths
    # ================================================================
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["F"].width = 2
    ws.column_dimensions["G"].width = 2
    ws.column_dimensions["H"].width = 38
    ws.column_dimensions["I"].width = 2
    ws.column_dimensions["J"].width = 16

    # ================================================================
    # Save
    # ================================================================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =========================================
# CSV EXPORT (V2)
# =========================================


def export_balance_sheet_to_csv(balance_sheet: BalanceSheet) -> str:
    """
    Export Balance Sheet to CSV - V2

    Args:
        balance_sheet: BalanceSheet object

    Returns:
        CSV content as string
    """
    output = StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(["BALANÇO GERENCIAL"])
    if balance_sheet.company_name:
        writer.writerow([balance_sheet.company_name])
    if balance_sheet.cnpj:
        writer.writerow([f"CNPJ: {balance_sheet.cnpj}"])
    writer.writerow(
        [f"Data de Referência: {balance_sheet.reference_date.strftime('%d/%m/%Y')}"]
    )
    writer.writerow([])

    year_str = str(balance_sheet.reference_date.year)

    # ================================================================
    # ATIVO
    # ================================================================
    writer.writerow(["ATIVO", year_str])
    writer.writerow(["Ativo Circulante", float(balance_sheet.ativo_circulante)])
    for line in balance_sheet.asset_lines:
        writer.writerow([f"  {line.name}", float(line.balance)])

    writer.writerow(["Ativo Não Circulante", float(balance_sheet.ativo_nao_circulante)])
    for line in balance_sheet.asset_noncurrent_lines:
        writer.writerow([f"  {line.name}", float(line.balance)])

    writer.writerow(["Imobilizado", float(balance_sheet.imobilizado)])
    for line in balance_sheet.imobilizado_lines:
        writer.writerow([f"  {line.name}", float(line.balance)])

    writer.writerow(["Intangível", float(balance_sheet.intangivel)])
    for line in balance_sheet.intangivel_lines:
        writer.writerow([f"  {line.name}", float(line.balance)])

    writer.writerow(["TOTAL DO ATIVO", float(balance_sheet.total_ativo)])
    writer.writerow([])

    # ================================================================
    # PASSIVO + PL
    # ================================================================
    writer.writerow(["PASSIVO + PL", year_str])

    writer.writerow(["Passivo Circulante", float(balance_sheet.passivo_circulante)])
    for line in balance_sheet.liability_lines:
        if line.code.startswith("2.01"):
            writer.writerow([f"  {line.name}", float(line.balance)])

    writer.writerow(["Passivo Não Circulante", float(balance_sheet.passivo_nao_circulante)])
    for line in balance_sheet.liability_lines + balance_sheet.liability_noncurrent_lines:
        if line.code.startswith("2.02"):
            writer.writerow([f"  {line.name}", float(line.balance)])

    writer.writerow(["Patrimônio Líquido", float(balance_sheet.patrimonio_liquido)])
    for line in balance_sheet.equity_lines:
        writer.writerow([f"  {line.name}", float(line.balance)])

    total_passivo_pl = balance_sheet.total_passivo + balance_sheet.patrimonio_liquido
    writer.writerow(["TOTAL DO PASSIVO + PL", float(total_passivo_pl)])
    writer.writerow([])

    # Balance check
    if balance_sheet.is_balanced:
        writer.writerow(["Balanço equilibrado (Ativo = Passivo + PL)"])
    else:
        writer.writerow(
            [f"Diferença: {format_brl(balance_sheet.balance_difference)}"]
        )

    output.seek(0)
    return output.getvalue()
