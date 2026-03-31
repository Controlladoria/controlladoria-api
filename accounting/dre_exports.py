"""
DRE Export Functions - V2 (Template-Based)
Export DRE to PDF, Excel, and CSV formats

V2 Changes:
- AV% calculated vs Receita Bruta (not Receita Líquida)
- New V2 structure: Variable Costs -> Contribution Margin -> Fixed Costs -> EBITDA
- Excel export matches ControlladorIA_Template_V0 layout
- Green color scheme matching template
"""

import io
from datetime import date
from decimal import Decimal
from typing import BinaryIO

from .dre_models import DRE, DRELine


def format_brl(value: Decimal) -> str:
    """
    Format decimal as Brazilian Real currency

    Args:
        value: Decimal amount

    Returns:
        Formatted string (e.g., "R$ 1.234,56")
    """
    # Handle negative values
    is_negative = value < 0
    abs_value = abs(value)

    # Format with thousands separator and 2 decimals
    formatted = f"{abs_value:,.2f}"

    # Replace separators (English → Brazilian)
    # English: 1,234.56 → Brazilian: 1.234,56
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")

    # Add currency symbol and negative formatting (parentheses for accounting)
    if is_negative:
        return f"(R$ {formatted})"
    else:
        return f"R$ {formatted}"


def format_percentage(value: float) -> str:
    """
    Format percentage with 1 decimal place

    Args:
        value: Percentage value

    Returns:
        Formatted string (e.g., "12,5%")
    """
    formatted = f"{value:.1f}%"
    return formatted.replace(".", ",")


# =========================================
# PDF EXPORT (V2)
# =========================================


def export_dre_to_pdf(dre: DRE, logo_bytes: bytes = None, prev_dre: "DRE | None" = None) -> bytes:
    """
    Export DRE to PDF format - V2 structure

    Args:
        dre: DRE object
        logo_bytes: Optional PNG/JPEG bytes for the header logo
        prev_dre: Optional previous period DRE for comparison

    Returns:
        PDF bytes
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (
            Image as RLImage,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        raise ImportError(
            "reportlab is required for PDF export. Install with: pip install reportlab"
        )

    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    # Container for PDF elements
    elements = []

    # Logo header
    if logo_bytes:
        try:
            img = RLImage(io.BytesIO(logo_bytes), width=5 * cm, height=2 * cm, kind="proportional")
            img.hAlign = "CENTER"
            elements.append(img)
            elements.append(Spacer(1, 0.3 * cm))
        except Exception:
            pass

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1a1a1a"),
        spaceAfter=30,
        alignment=1,  # Center
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#666666"),
        spaceAfter=20,
        alignment=1,  # Center
    )

    # Header
    company_name = dre.company_name or "Empresa"
    elements.append(Paragraph("DRE GERENCIAL", title_style))
    elements.append(Paragraph(f"<b>{company_name}</b>", subtitle_style))

    if dre.cnpj:
        elements.append(Paragraph(f"CNPJ: {dre.cnpj}", subtitle_style))

    period_str = (
        f"{dre.start_date.strftime('%d/%m/%Y')} a {dre.end_date.strftime('%d/%m/%Y')}"
    )
    elements.append(Paragraph(f"Período: {period_str}", subtitle_style))
    elements.append(Spacer(1, 0.5 * cm))

    # DRE Table Data
    table_data = []
    has_prev = prev_dre is not None

    # Table header
    if has_prev:
        prev_label = str(prev_dre.end_date.year) if prev_dre.end_date.year != dre.end_date.year else f"{prev_dre.start_date.strftime('%m/%Y')}"
        curr_label = str(dre.end_date.year) if prev_dre.end_date.year != dre.end_date.year else f"{dre.start_date.strftime('%m/%Y')}"
        table_data.append(
            [
                Paragraph("<b>Descrição</b>", styles["Normal"]),
                Paragraph(f"<b>{prev_label}</b>", styles["Normal"]),
                Paragraph("<b>AV%</b>", styles["Normal"]),
                Paragraph(f"<b>{curr_label}</b>", styles["Normal"]),
                Paragraph("<b>AV%</b>", styles["Normal"]),
            ]
        )
    else:
        table_data.append(
            [
                Paragraph("<b>Descrição</b>", styles["Normal"]),
                Paragraph(f"<b>{dre.end_date.year}</b>", styles["Normal"]),
                Paragraph("<b>AV%</b>", styles["Normal"]),
            ]
        )

    # V2: AV% calculated vs Receita Bruta
    gross_revenue_float = float(dre.receita_bruta)

    def av_pct(amount: Decimal) -> float:
        if gross_revenue_float == 0:
            return 0
        return (float(amount) / gross_revenue_float) * 100

    prev_gross_revenue_float = float(prev_dre.receita_bruta) if has_prev else 0

    def prev_av_pct(amount: Decimal) -> float:
        if prev_gross_revenue_float == 0:
            return 0
        return (float(amount) / prev_gross_revenue_float) * 100

    # Helper function to add lines
    def add_line(
        description: str,
        value: Decimal,
        pct_val: float = None,
        is_total: bool = False,
        level: int = 0,
        prev_value: Decimal = None,
        prev_pct: float = None,
    ):
        indent = "&nbsp;" * level * 4

        if is_total:
            desc = f"<b>{indent}{description}</b>"
            val = f"<b>{format_brl(value)}</b>"
            pct_str = f"<b>{format_percentage(pct_val)}</b>" if pct_val is not None else ""
        else:
            desc = f"{indent}{description}"
            val = format_brl(value)
            pct_str = format_percentage(pct_val) if pct_val is not None else ""

        if has_prev:
            if is_total:
                pval = f"<b>{format_brl(prev_value)}</b>" if prev_value is not None else "<b>R$ 0,00</b>"
                ppct = f"<b>{format_percentage(prev_pct)}</b>" if prev_pct is not None else ""
            else:
                pval = format_brl(prev_value) if prev_value is not None else "R$ 0,00"
                ppct = format_percentage(prev_pct) if prev_pct is not None else ""
            table_data.append(
                [
                    Paragraph(desc, styles["Normal"]),
                    Paragraph(pval, styles["Normal"]),
                    Paragraph(ppct, styles["Normal"]),
                    Paragraph(val, styles["Normal"]),
                    Paragraph(pct_str, styles["Normal"]),
                ]
            )
        else:
            table_data.append(
                [
                    Paragraph(desc, styles["Normal"]),
                    Paragraph(val, styles["Normal"]),
                    Paragraph(pct_str, styles["Normal"]),
                ]
            )

    # ================================================================
    # V2 DRE Lines (matching ControlladorIA template)
    # ================================================================
    p = prev_dre if has_prev else None

    # Helper: find a sub-item amount in prev_dre by code
    def _prev_sub_amount(code: str) -> Decimal:
        if not has_prev or not prev_dre.detailed_lines:
            return Decimal("0")
        for pl in prev_dre.detailed_lines:
            if pl.code == code:
                return pl.amount
        return Decimal("0")

    # Helper: render individual items from detailed_lines by code prefix
    def _render_sub_items(code_prefix: str, level: int = 1):
        for line in dre.detailed_lines:
            if (line.level >= 2 and not line.is_subtotal and not line.is_total
                    and line.code.startswith(code_prefix) and line.amount != 0):
                pv = _prev_sub_amount(line.code)
                add_line(line.description, line.amount, av_pct(abs(line.amount)), level=level,
                         prev_value=pv, prev_pct=prev_av_pct(abs(pv)))

    # Receita Bruta
    add_line("Receita Bruta", dre.receita_bruta, 100.0, is_total=True,
             prev_value=p.receita_bruta if p else None, prev_pct=100.0 if p else None)
    _render_sub_items("1.1")

    # (-) Deduções
    if dre.total_deducoes > 0 or (p and p.total_deducoes > 0):
        add_line("(-) Deduções e Impostos sobre Vendas", -dre.total_deducoes, -av_pct(dre.total_deducoes),
                 prev_value=-p.total_deducoes if p else None, prev_pct=-prev_av_pct(p.total_deducoes) if p else None)
        _render_sub_items("1.2")

    # Receita Líquida
    add_line("Receita Líquida", dre.receita_liquida, av_pct(dre.receita_liquida), is_total=True,
             prev_value=p.receita_liquida if p else None, prev_pct=prev_av_pct(p.receita_liquida) if p else None)

    # Custos Variáveis
    if dre.total_custos_variaveis > 0 or (p and p.total_custos_variaveis > 0):
        add_line("Custos Variáveis", -dre.total_custos_variaveis, -av_pct(dre.total_custos_variaveis),
                 prev_value=-p.total_custos_variaveis if p else None, prev_pct=-prev_av_pct(p.total_custos_variaveis) if p else None)
        _render_sub_items("2.")

    # Margem de Contribuição
    add_line("Margem de Contribuição", dre.margem_contribuicao, av_pct(dre.margem_contribuicao), is_total=True,
             prev_value=p.margem_contribuicao if p else None, prev_pct=prev_av_pct(p.margem_contribuicao) if p else None)

    # Custos Fixos + Despesas Fixas
    if dre.despesas_administrativas > 0 or (p and p.despesas_administrativas > 0):
        add_line("Despesas Administrativas", -dre.despesas_administrativas, -av_pct(dre.despesas_administrativas),
                 prev_value=-p.despesas_administrativas if p else None, prev_pct=-prev_av_pct(p.despesas_administrativas) if p else None)
        _render_sub_items("3.1", level=1)
    if dre.despesas_vendas > 0 or (p and p.despesas_vendas > 0):
        add_line("Despesas Comerciais", -dre.despesas_vendas, -av_pct(dre.despesas_vendas),
                 prev_value=-p.despesas_vendas if p else None, prev_pct=-prev_av_pct(p.despesas_vendas) if p else None)
        _render_sub_items("3.2", level=1)
    if dre.custos_fixos_producao > 0 or (p and p.custos_fixos_producao > 0):
        add_line("Custos Fixos Operacionais", -dre.custos_fixos_producao, -av_pct(dre.custos_fixos_producao),
                 prev_value=-p.custos_fixos_producao if p else None, prev_pct=-prev_av_pct(p.custos_fixos_producao) if p else None)
    if dre.outras_despesas > 0 or (p and p.outras_despesas > 0):
        add_line("Outras Despesas", -dre.outras_despesas, -av_pct(dre.outras_despesas),
                 prev_value=-p.outras_despesas if p else None, prev_pct=-prev_av_pct(p.outras_despesas) if p else None)

    # EBITDA
    add_line("EBITDA", dre.ebitda, av_pct(dre.ebitda), is_total=True,
             prev_value=p.ebitda if p else None, prev_pct=prev_av_pct(p.ebitda) if p else None)

    # D&A
    if dre.total_deprec_amort > 0 or (p and p.total_deprec_amort > 0):
        add_line("Depreciação e Amortização", -dre.total_deprec_amort, -av_pct(dre.total_deprec_amort),
                 prev_value=-p.total_deprec_amort if p else None, prev_pct=-prev_av_pct(p.total_deprec_amort) if p else None)

    # Lucro Operacional
    add_line("Lucro Operacional", dre.resultado_operacional, av_pct(dre.resultado_operacional), is_total=True,
             prev_value=p.resultado_operacional if p else None, prev_pct=prev_av_pct(p.resultado_operacional) if p else None)

    # Resultado Financeiro
    add_line("Resultado Financeiro (líquido)", dre.resultado_financeiro, av_pct(dre.resultado_financeiro),
             prev_value=p.resultado_financeiro if p else None, prev_pct=prev_av_pct(p.resultado_financeiro) if p else None)
    _render_sub_items("1.4")
    _render_sub_items("4.")

    # LAIR
    add_line("Lucro Antes do IR/CSLL", dre.resultado_antes_impostos, av_pct(dre.resultado_antes_impostos), is_total=True,
             prev_value=p.resultado_antes_impostos if p else None, prev_pct=prev_av_pct(p.resultado_antes_impostos) if p else None)

    # Impostos
    if dre.total_impostos_lucro > 0 or (p and p.total_impostos_lucro > 0):
        add_line("IRPJ e CSLL", -dre.total_impostos_lucro, -av_pct(dre.total_impostos_lucro),
                 prev_value=-p.total_impostos_lucro if p else None, prev_pct=-prev_av_pct(p.total_impostos_lucro) if p else None)

    # Lucro Líquido
    add_line("Lucro Líquido", dre.lucro_liquido, av_pct(dre.lucro_liquido), is_total=True,
             prev_value=p.lucro_liquido if p else None, prev_pct=prev_av_pct(p.lucro_liquido) if p else None)

    # Create table
    if has_prev:
        table = Table(table_data, colWidths=[8 * cm, 3 * cm, 2.5 * cm, 3 * cm, 2.5 * cm])
    else:
        table = Table(table_data, colWidths=[11 * cm, 3.5 * cm, 3 * cm])

    # Table styling - green theme matching template
    table.setStyle(
        TableStyle(
            [
                # Header row
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                # Body
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),  # Value and % columns
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                # Grid
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("LINEBELOW", (0, 0), (-1, 0), 2, colors.HexColor("#388E3C")),
                # Alternating row colors
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f1f8e9")],
                ),
            ]
        )
    )

    elements.append(table)

    # Build PDF
    doc.build(elements)

    # Get PDF bytes
    pdf_bytes = buffer.getvalue()
    buffer.close()

    return pdf_bytes


# =========================================
# EXCEL EXPORT (V2 - Template-based)
# =========================================


def export_dre_to_excel(dre: DRE, logo_bytes: bytes = None, prev_dre: "DRE | None" = None) -> bytes:
    """
    Export DRE to Excel format matching ControlladorIA_Template_V0 layout.

    Layout:
    - B2: "DRE GERENCIAL"
    - When prev_dre is provided: D4: Prev Year, E4: Prev AV%, F4: Current Year, G4: Current AV%
    - Otherwise: D4: Year, E4: "AV%"
    - Green color scheme
    - Total/subtotal rows highlighted

    Args:
        dre: DRE object (current period)
        logo_bytes: Optional PNG/JPEG bytes for the header logo
        prev_dre: Optional previous period DRE for comparison

    Returns:
        Excel file bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    # Create workbook
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "DRE GERENCIAL"

    # Logo header (floats over A1)
    if logo_bytes:
        try:
            from openpyxl.drawing.image import Image as XLImage
            xl_img = XLImage(io.BytesIO(logo_bytes))
            xl_img.width = 140
            xl_img.height = 50
            ws.add_image(xl_img, "A1")
        except Exception:
            pass

    # ================================================================
    # Styles matching ControlladorIA template
    # ================================================================
    green_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    light_green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    title_font = Font(name="Arial", size=14, bold=True, color="1B5E20")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    total_font = Font(name="Arial", size=11, bold=True)
    normal_font = Font(name="Arial", size=10)
    sub_font = Font(name="Arial", size=10, color="333333")

    thin_border = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD"),
    )

    # ================================================================
    # Header
    # ================================================================
    has_prev = prev_dre is not None
    merge_end = "G2" if has_prev else "E2"

    ws["B2"] = "DRE GERENCIAL"
    ws["B2"].font = title_font
    ws.merge_cells(f"B2:{merge_end}")

    company_name = dre.company_name or "Empresa"
    ws["B3"] = company_name
    ws["B3"].font = Font(name="Arial", size=11, color="666666")

    # Column headers - layout depends on whether we have previous period
    if has_prev:
        # Previous period columns: D, E | Current period columns: F, G
        prev_label = str(prev_dre.end_date.year) if prev_dre.end_date.year != dre.end_date.year else f"{prev_dre.start_date.strftime('%m/%Y')}"
        curr_label = str(dre.end_date.year) if prev_dre.end_date.year != dre.end_date.year else f"{dre.start_date.strftime('%m/%Y')}"

        for col, label in [("D", prev_label), ("E", "AV%"), ("F", curr_label), ("G", "AV%")]:
            ws[f"{col}4"] = label
            ws[f"{col}4"].font = header_font
            ws[f"{col}4"].fill = green_fill
            ws[f"{col}4"].alignment = Alignment(horizontal="center")
            ws[f"{col}4"].border = thin_border
    else:
        year_str = str(dre.end_date.year)
        ws["D4"] = year_str
        ws["D4"].font = header_font
        ws["D4"].fill = green_fill
        ws["D4"].alignment = Alignment(horizontal="center")
        ws["D4"].border = thin_border

        ws["E4"] = "AV%"
        ws["E4"].font = header_font
        ws["E4"].fill = green_fill
        ws["E4"].alignment = Alignment(horizontal="center")
        ws["E4"].border = thin_border

    # ================================================================
    # AV% helpers
    # ================================================================
    gross_revenue_float = float(dre.receita_bruta)

    def av_pct(amount: Decimal) -> float:
        """AV% for Excel: returns 0-1 range (Excel '0.00%' format multiplies by 100)"""
        if gross_revenue_float == 0:
            return 0
        return float(amount) / gross_revenue_float

    prev_gross_revenue_float = float(prev_dre.receita_bruta) if has_prev else 0

    def prev_av_pct(amount: Decimal) -> float:
        """AV% for previous period: returns 0-1 range"""
        if prev_gross_revenue_float == 0:
            return 0
        return float(amount) / prev_gross_revenue_float

    # ================================================================
    # DRE Lines (matching template rows B5-B31)
    # ================================================================
    row = 5

    def write_line(description: str, value: Decimal, pct_val: float = None,
                   is_total: bool = False, is_highlight: bool = False,
                   level: int = 0,
                   prev_value: Decimal = None, prev_pct: float = None):
        nonlocal row
        # Description with indent
        ws[f"B{row}"] = ("  " * level) + description

        if has_prev:
            # Previous period in D, E
            ws[f"D{row}"] = float(prev_value) if prev_value is not None else 0
            if prev_pct is not None:
                ws[f"E{row}"] = prev_pct
            ws[f"D{row}"].number_format = '#,##0.00'
            ws[f"E{row}"].number_format = '0.00%'
            ws[f"D{row}"].alignment = Alignment(horizontal="right")
            ws[f"E{row}"].alignment = Alignment(horizontal="right")

            # Current period in F, G
            ws[f"F{row}"] = float(value)
            if pct_val is not None:
                ws[f"G{row}"] = pct_val
            ws[f"F{row}"].number_format = '#,##0.00'
            ws[f"G{row}"].number_format = '0.00%'
            ws[f"F{row}"].alignment = Alignment(horizontal="right")
            ws[f"G{row}"].alignment = Alignment(horizontal="right")

            data_cols = ["B", "D", "E", "F", "G"]
        else:
            ws[f"D{row}"] = float(value)
            if pct_val is not None:
                ws[f"E{row}"] = pct_val
            ws[f"D{row}"].number_format = '#,##0.00'
            ws[f"E{row}"].number_format = '0.00%'
            ws[f"D{row}"].alignment = Alignment(horizontal="right")
            ws[f"E{row}"].alignment = Alignment(horizontal="right")
            data_cols = ["B", "D", "E"]

        if is_total:
            for col in data_cols:
                ws[f"{col}{row}"].font = total_font
            fill = light_green_fill if is_highlight else subtotal_fill
            all_cols = ["B", "C", "D", "E", "F", "G"] if has_prev else ["B", "C", "D", "E"]
            for col in all_cols:
                ws[f"{col}{row}"].fill = fill
        else:
            ws[f"B{row}"].font = sub_font if level > 0 else normal_font
            for col in data_cols:
                if col != "B":
                    ws[f"{col}{row}"].font = normal_font

        for col in data_cols:
            ws[f"{col}{row}"].border = thin_border

        row += 1

    # --- DRE BODY (matching template exactly) ---

    # Helper: find a sub-item amount in prev_dre by code
    def _prev_sub_amount(code: str) -> Decimal:
        if not has_prev or not prev_dre.detailed_lines:
            return Decimal("0")
        for pl in prev_dre.detailed_lines:
            if pl.code == code:
                return pl.amount
        return Decimal("0")

    # Helper: render individual items from detailed_lines by code prefix
    def _xl_sub_items(code_prefix: str, level: int = 1):
        for line in dre.detailed_lines:
            if (line.level >= 2 and not line.is_subtotal and not line.is_total
                    and line.code.startswith(code_prefix) and line.amount != 0):
                pv = _prev_sub_amount(line.code)
                write_line(line.description, line.amount, av_pct(abs(line.amount)), level=level,
                           prev_value=pv, prev_pct=prev_av_pct(abs(pv)))

    # Shorthand for previous period values
    p = prev_dre if has_prev else None

    # Receita Bruta
    write_line("Receita Bruta", dre.receita_bruta, 1.0, is_total=True, is_highlight=True,
               prev_value=p.receita_bruta if p else None, prev_pct=1.0 if p else None)
    _xl_sub_items("1.1")  # Revenue sub-items (1.1.XX)

    # (-) Deduções
    write_line("(-) Deduções e Impostos sobre Vendas", -dre.total_deducoes, -av_pct(dre.total_deducoes),
               prev_value=-p.total_deducoes if p else None, prev_pct=-prev_av_pct(p.total_deducoes) if p else None)
    _xl_sub_items("1.2")  # Deduction sub-items (1.2.XX)

    # Receita Líquida
    write_line("Receita Líquida", dre.receita_liquida, av_pct(dre.receita_liquida), is_total=True,
               prev_value=p.receita_liquida if p else None, prev_pct=prev_av_pct(p.receita_liquida) if p else None)

    # Custos Variáveis
    write_line("Custos Variáveis", -dre.total_custos_variaveis, -av_pct(dre.total_custos_variaveis),
               prev_value=-p.total_custos_variaveis if p else None, prev_pct=-prev_av_pct(p.total_custos_variaveis) if p else None)
    _xl_sub_items("2.")  # Variable cost sub-items (2.X.XX)

    # Margem de Contribuição
    write_line("Margem de Contribuição", dre.margem_contribuicao, av_pct(dre.margem_contribuicao),
               is_total=True, is_highlight=True,
               prev_value=p.margem_contribuicao if p else None, prev_pct=prev_av_pct(p.margem_contribuicao) if p else None)

    # Custos Fixos + Despesas Fixas
    if dre.despesas_administrativas > 0 or (p and p.despesas_administrativas > 0):
        write_line("Despesas Administrativas", -dre.despesas_administrativas, -av_pct(dre.despesas_administrativas),
                   prev_value=-p.despesas_administrativas if p else None, prev_pct=-prev_av_pct(p.despesas_administrativas) if p else None)
        _xl_sub_items("3.1", level=1)  # Admin expense sub-items (3.1.XX)
    if dre.despesas_vendas > 0 or (p and p.despesas_vendas > 0):
        write_line("Despesas Comerciais", -dre.despesas_vendas, -av_pct(dre.despesas_vendas),
                   prev_value=-p.despesas_vendas if p else None, prev_pct=-prev_av_pct(p.despesas_vendas) if p else None)
        _xl_sub_items("3.2", level=1)  # Commercial expense sub-items (3.2.XX)
    if dre.custos_fixos_producao > 0 or (p and p.custos_fixos_producao > 0):
        write_line("Custos Fixos Operacionais", -dre.custos_fixos_producao, -av_pct(dre.custos_fixos_producao),
                   prev_value=-p.custos_fixos_producao if p else None, prev_pct=-prev_av_pct(p.custos_fixos_producao) if p else None)
    if dre.outras_despesas > 0 or (p and p.outras_despesas > 0):
        write_line("Outras Despesas", -dre.outras_despesas, -av_pct(dre.outras_despesas),
                   prev_value=-p.outras_despesas if p else None, prev_pct=-prev_av_pct(p.outras_despesas) if p else None)

    # EBITDA
    write_line("EBITDA", dre.ebitda, av_pct(dre.ebitda), is_total=True, is_highlight=True,
               prev_value=p.ebitda if p else None, prev_pct=prev_av_pct(p.ebitda) if p else None)

    # Depreciação e Amortização
    write_line("Depreciação e Amortização", -dre.total_deprec_amort, -av_pct(dre.total_deprec_amort),
               prev_value=-p.total_deprec_amort if p else None, prev_pct=-prev_av_pct(p.total_deprec_amort) if p else None)

    # Lucro Operacional
    write_line("Lucro Operacional", dre.resultado_operacional, av_pct(dre.resultado_operacional), is_total=True,
               prev_value=p.resultado_operacional if p else None, prev_pct=prev_av_pct(p.resultado_operacional) if p else None)

    # Resultado Financeiro
    write_line("Resultado Financeiro (líquido)", dre.resultado_financeiro, av_pct(dre.resultado_financeiro),
               prev_value=p.resultado_financeiro if p else None, prev_pct=prev_av_pct(p.resultado_financeiro) if p else None)
    _xl_sub_items("1.4")  # Financial/other revenue sub-items
    _xl_sub_items("4.")   # Financial expense sub-items

    # Lucro Antes do IR/CSLL
    write_line("Lucro Antes do IR/CSLL", dre.resultado_antes_impostos, av_pct(dre.resultado_antes_impostos), is_total=True,
               prev_value=p.resultado_antes_impostos if p else None, prev_pct=prev_av_pct(p.resultado_antes_impostos) if p else None)

    # IRPJ e CSLL
    write_line("IRPJ e CSLL", -dre.total_impostos_lucro, -av_pct(dre.total_impostos_lucro),
               prev_value=-p.total_impostos_lucro if p else None, prev_pct=-prev_av_pct(p.total_impostos_lucro) if p else None)

    # Lucro Líquido
    write_line("Lucro Líquido", dre.lucro_liquido, av_pct(dre.lucro_liquido),
               is_total=True, is_highlight=True,
               prev_value=p.lucro_liquido if p else None, prev_pct=prev_av_pct(p.lucro_liquido) if p else None)

    # ================================================================
    # Column widths
    # ================================================================
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    if has_prev:
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 12

    # ================================================================
    # Save
    # ================================================================
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()

    return excel_bytes


# =========================================
# CSV EXPORT (V2)
# =========================================


def export_dre_to_csv(dre: DRE) -> str:
    """
    Export DRE to CSV format - V2 structure matching template

    Args:
        dre: DRE object

    Returns:
        CSV string
    """
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    company_name = dre.company_name or "Empresa"
    period_str = (
        f"{dre.start_date.strftime('%d/%m/%Y')} a {dre.end_date.strftime('%d/%m/%Y')}"
    )

    writer.writerow(["DRE GERENCIAL"])
    writer.writerow([company_name])
    if dre.cnpj:
        writer.writerow([f"CNPJ: {dre.cnpj}"])
    writer.writerow([f"Período: {period_str}"])
    writer.writerow([])  # Empty row

    # Column headers
    writer.writerow(["Descrição", f"{dre.end_date.year}", "AV%"])

    # V2: AV% calculated vs Receita Bruta
    gross_revenue_float = float(dre.receita_bruta)

    def av_pct(amount: Decimal) -> float:
        if gross_revenue_float == 0:
            return 0
        return (float(amount) / gross_revenue_float) * 100

    # Helper to add row
    def add_csv_row(description: str, value: Decimal, pct_val: float = None):
        pct_str = f"{pct_val:.1f}%".replace(".", ",") if pct_val is not None else ""
        value_str = f"{float(value):.2f}".replace(".", ",")
        writer.writerow([description, value_str, pct_str])

    # ================================================================
    # DRE Lines (matching template)
    # ================================================================

    # Helper: render individual items from detailed_lines by code prefix
    def _csv_sub_items(code_prefix: str, indent: str = "  "):
        for line in dre.detailed_lines:
            if (line.level >= 2 and not line.is_subtotal and not line.is_total
                    and line.code.startswith(code_prefix) and line.amount != 0):
                add_csv_row(f"{indent}{line.description}", line.amount, av_pct(abs(line.amount)))

    add_csv_row("Receita Bruta", dre.receita_bruta, 100.0)
    _csv_sub_items("1.1")  # Revenue sub-items (1.1.XX)

    add_csv_row("(-) Deduções e Impostos sobre Vendas", -dre.total_deducoes, -av_pct(dre.total_deducoes))
    _csv_sub_items("1.2")  # Deduction sub-items (1.2.XX)

    add_csv_row("Receita Líquida", dre.receita_liquida, av_pct(dre.receita_liquida))

    add_csv_row("Custos Variáveis", -dre.total_custos_variaveis, -av_pct(dre.total_custos_variaveis))
    _csv_sub_items("2.")  # Variable cost sub-items (2.X.XX)

    add_csv_row("Margem de Contribuição", dre.margem_contribuicao, av_pct(dre.margem_contribuicao))

    if dre.despesas_administrativas > 0:
        add_csv_row("Despesas Administrativas", -dre.despesas_administrativas, -av_pct(dre.despesas_administrativas))
        _csv_sub_items("3.1", indent="    ")
    if dre.despesas_vendas > 0:
        add_csv_row("Despesas Comerciais", -dre.despesas_vendas, -av_pct(dre.despesas_vendas))
        _csv_sub_items("3.2", indent="    ")
    if dre.custos_fixos_producao > 0:
        add_csv_row("Custos Fixos Operacionais", -dre.custos_fixos_producao, -av_pct(dre.custos_fixos_producao))
    if dre.outras_despesas > 0:
        add_csv_row("Outras Despesas", -dre.outras_despesas, -av_pct(dre.outras_despesas))

    add_csv_row("EBITDA", dre.ebitda, av_pct(dre.ebitda))

    add_csv_row("Depreciação e Amortização", -dre.total_deprec_amort, -av_pct(dre.total_deprec_amort))

    add_csv_row("Lucro Operacional", dre.resultado_operacional, av_pct(dre.resultado_operacional))

    add_csv_row("Resultado Financeiro (líquido)", dre.resultado_financeiro, av_pct(dre.resultado_financeiro))
    _csv_sub_items("1.4")  # Financial/other revenue sub-items
    _csv_sub_items("4.")   # Financial expense sub-items

    add_csv_row("Lucro Antes do IR/CSLL", dre.resultado_antes_impostos, av_pct(dre.resultado_antes_impostos))

    add_csv_row("IRPJ e CSLL", -dre.total_impostos_lucro, -av_pct(dre.total_impostos_lucro))

    add_csv_row("Lucro Líquido", dre.lucro_liquido, av_pct(dre.lucro_liquido))

    csv_string = output.getvalue()
    output.close()

    return csv_string
