"""
Cash Flow export functions
Generates PDF, Excel, and CSV reports for Fluxo de Caixa

Structure: Saldo Inicial -> Entradas (by category) -> Saídas (by category) -> Saldo Final
Green color scheme matching ControlladorIA template.
"""

from decimal import Decimal
from io import BytesIO
from typing import Optional

from accounting.cash_flow_calculator import CashFlow


def _format_currency(value: Decimal) -> str:
    """Format decimal as Brazilian Real currency"""
    abs_value = abs(value)
    formatted = f"{abs_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if value < 0:
        return f"-R$ {formatted}"
    return f"R$ {formatted}"


def export_cash_flow_to_pdf(cash_flow: CashFlow, logo_bytes: bytes = None) -> bytes:
    """Export Cash Flow to PDF — Saldo Inicial / Entradas / Saídas / Saldo Final"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        Image as RLImage,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=10 * mm, bottomMargin=10 * mm)
    story = []
    styles = getSampleStyleSheet()

    if logo_bytes:
        try:
            img = RLImage(BytesIO(logo_bytes), width=5 * cm, height=2 * cm, kind="proportional")
            img.hAlign = "CENTER"
            story.append(img)
            story.append(Spacer(1, 3 * mm))
        except Exception:
            pass

    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Heading1"],
        fontSize=16, textColor=colors.HexColor("#1B5E20"), spaceAfter=20, alignment=1,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle", parent=styles["Normal"],
        fontSize=10, textColor=colors.HexColor("#666666"), spaceAfter=10, alignment=1,
    )

    story.append(Paragraph("FLUXO DE CAIXA", title_style))
    story.append(Paragraph(cash_flow.company_name, subtitle_style))
    if cash_flow.cnpj:
        story.append(Paragraph(f"CNPJ: {cash_flow.cnpj}", subtitle_style))
    story.append(Paragraph(
        f"Período: {cash_flow.start_date.strftime('%d/%m/%Y')} a {cash_flow.end_date.strftime('%d/%m/%Y')}",
        subtitle_style,
    ))
    story.append(Spacer(1, 15))

    # Build table
    table_data = []
    table_data.append([
        Paragraph("<b>Descrição</b>", styles["Normal"]),
        Paragraph(f"<b>Valor</b>", styles["Normal"]),
    ])

    # SALDO INICIAL
    table_data.append(["SALDO INICIAL", _format_currency(cash_flow.cash_beginning)])

    # ENTRADAS
    table_data.append(["", ""])
    table_data.append(["ENTRADAS", _format_currency(cash_flow.operating_activities.total)])
    for label, value in cash_flow.operating_activities.line_items.items():
        table_data.append([f"  (+) {label}", _format_currency(value)])

    # SAÍDAS
    table_data.append(["", ""])
    table_data.append(["SAÍDAS", _format_currency(cash_flow.investing_activities.total)])
    for label, value in cash_flow.investing_activities.line_items.items():
        table_data.append([f"  (-) {label}", _format_currency(value)])

    # RESULTADO
    table_data.append(["", ""])
    table_data.append(["VARIAÇÃO NO PERÍODO", _format_currency(cash_flow.net_increase_in_cash)])
    table_data.append(["SALDO FINAL", _format_currency(cash_flow.cash_ending)])

    table = Table(table_data, colWidths=[130 * mm, 50 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f8e9")]),
    ]))

    for i, row_data in enumerate(table_data):
        if row_data[0] in ["SALDO INICIAL", "SALDO FINAL", "ENTRADAS", "SAÍDAS", "VARIAÇÃO NO PERÍODO"]:
            table.setStyle(TableStyle([
                ("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"),
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#C8E6C9")),
            ]))

    story.append(table)
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def export_cash_flow_to_excel(cash_flow: CashFlow, logo_bytes: bytes = None) -> bytes:
    """Export Cash Flow to Excel — Saldo Inicial / Entradas / Saídas / Saldo Final"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "FLUXO DE CAIXA"

    if logo_bytes:
        try:
            from openpyxl.drawing.image import Image as XLImage
            xl_img = XLImage(BytesIO(logo_bytes))
            xl_img.width = 140
            xl_img.height = 50
            ws.add_image(xl_img, "A1")
        except Exception:
            pass

    green_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    light_green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    title_font = Font(name="Arial", size=14, bold=True, color="1B5E20")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Arial", size=11, bold=True)
    total_font = Font(name="Arial", size=11, bold=True)
    normal_font = Font(name="Arial", size=10)
    sub_font = Font(name="Arial", size=10, color="333333")
    thin_border = Border(
        left=Side(style="thin", color="BDBDBD"), right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"), bottom=Side(style="thin", color="BDBDBD"),
    )

    period = f"{cash_flow.start_date.strftime('%d/%m/%Y')} a {cash_flow.end_date.strftime('%d/%m/%Y')}"
    ws["B2"] = f"FLUXO DE CAIXA"
    ws["B2"].font = title_font
    ws.merge_cells("B2:D2")
    ws["B3"] = cash_flow.company_name
    ws["B3"].font = Font(name="Arial", size=11, color="666666")
    ws["B4"] = f"Período: {period}"
    ws["B4"].font = Font(name="Arial", size=10, color="999999")

    ws["D5"] = "Valor"
    ws["D5"].font = header_font
    ws["D5"].fill = green_fill
    ws["D5"].alignment = Alignment(horizontal="center")
    ws["D5"].border = thin_border

    row = 6

    def write_section_header(label):
        nonlocal row
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = section_font
        ws[f"B{row}"].fill = light_green_fill
        ws[f"D{row}"].fill = light_green_fill
        for col in ["B", "D"]:
            ws[f"{col}{row}"].border = thin_border
        row += 1

    def write_line(label, value, is_total=False, level=0):
        nonlocal row
        ws[f"B{row}"] = ("  " * level) + label
        ws[f"D{row}"] = float(value)
        ws[f"D{row}"].number_format = '#,##0.00'
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        if is_total:
            ws[f"B{row}"].font = total_font
            ws[f"D{row}"].font = total_font
            for col in ["B", "D"]:
                ws[f"{col}{row}"].fill = subtotal_fill
        else:
            ws[f"B{row}"].font = sub_font if level > 0 else normal_font
            ws[f"D{row}"].font = normal_font
        for col in ["B", "D"]:
            ws[f"{col}{row}"].border = thin_border
        row += 1

    # SALDO INICIAL
    write_section_header("SALDO INICIAL")
    write_line("Saldo de Caixa", cash_flow.cash_beginning, is_total=True)
    row += 1

    # ENTRADAS
    write_section_header("ENTRADAS")
    for label, value in cash_flow.operating_activities.line_items.items():
        write_line(f"(+) {label}", value, level=1)
    write_line("Total Entradas", cash_flow.operating_activities.total, is_total=True)
    row += 1

    # SAÍDAS
    write_section_header("SAÍDAS")
    for label, value in cash_flow.investing_activities.line_items.items():
        write_line(f"(-) {label}", value, level=1)
    write_line("Total Saídas", cash_flow.investing_activities.total, is_total=True)
    row += 1

    # RESULTADO
    write_section_header("RESULTADO")
    write_line("Variação no Período", cash_flow.net_increase_in_cash, is_total=True)
    write_line("Saldo Final", cash_flow.cash_ending, is_total=True)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes


def export_cash_flow_to_csv(cash_flow: CashFlow) -> str:
    """Export Cash Flow to CSV — Saldo Inicial / Entradas / Saídas / Saldo Final"""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(["FLUXO DE CAIXA"])
    writer.writerow([cash_flow.company_name])
    if cash_flow.cnpj:
        writer.writerow([f"CNPJ: {cash_flow.cnpj}"])
    writer.writerow([
        f"Período: {cash_flow.start_date.strftime('%d/%m/%Y')} a {cash_flow.end_date.strftime('%d/%m/%Y')}"
    ])
    writer.writerow([])
    writer.writerow(["Descrição", "Valor"])

    # SALDO INICIAL
    writer.writerow(["SALDO INICIAL", f"{float(cash_flow.cash_beginning):.2f}"])
    writer.writerow([])

    # ENTRADAS
    writer.writerow(["ENTRADAS", f"{float(cash_flow.operating_activities.total):.2f}"])
    for label, value in cash_flow.operating_activities.line_items.items():
        writer.writerow([f"  (+) {label}", f"{float(value):.2f}"])
    writer.writerow([])

    # SAÍDAS
    writer.writerow(["SAÍDAS", f"{float(cash_flow.investing_activities.total):.2f}"])
    for label, value in cash_flow.investing_activities.line_items.items():
        writer.writerow([f"  (-) {label}", f"{float(value):.2f}"])
    writer.writerow([])

    # RESULTADO
    writer.writerow(["VARIAÇÃO NO PERÍODO", f"{float(cash_flow.net_increase_in_cash):.2f}"])
    writer.writerow(["SALDO FINAL", f"{float(cash_flow.cash_ending):.2f}"])

    csv_string = output.getvalue()
    output.close()
    return csv_string
